from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.cache import never_cache

from .models import Switch, Port
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from django.contrib.auth.decorators import login_required

@csrf_exempt
@never_cache
@login_required(login_url="/login")
def index(request):
    switches = Switch.objects.all()
    return render(request, "index.html", {"switches": switches})


@login_required(login_url='login')
def switch_detail(request, pk):
    switch = get_object_or_404(Switch, pk=pk)
    ports = switch.ports.all()

    port_type_choices = dict(Port._meta.get_field('portType').choices)
    status_choices = dict(Port._meta.get_field('status').choices)

    # Возвращаем HTML вместо JSON
    return render(request, "switch_detail.html", {
        "switch": switch,
        "ports": ports,
        "port_type_choices": port_type_choices,
        "status_choices": status_choices
    })

@csrf_exempt
def update_ports(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            for item in data:
                port = Port.objects.get(id=item["id"])
                port.connected_device = item.get("connected_device") or None
                port.ip_address = item.get("ip_address") or None
                port.device_name = item.get("device_name") or None
                port.portType = item.get("portType") or None
                port.status = item.get("status") or None
                port.vlanNum = item.get("vlanNum") or None
                port.upLink = item.get("upLink") or None
                port.addInfo = item.get("addInfo") or None
                port.cabPlacement = item.get("cabPlacement") or None
                port.save()
            return JsonResponse({"status": "ok"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error"}, status=400)

@csrf_exempt
@require_POST
def update_port(request, pk):
    try:
        data = json.loads(request.body)
        port = Port.objects.get(pk=pk)

        port.connected_device = data.get("connected_device") or None
        port.ip_address = data.get("ip_address") or None
        port.device_name = data.get("device_name") or None
        port.portType = data.get("portType") or None
        port.status = data.get("status") or None
        port.vlanNum = data.get("vlanNum") or None
        port.upLink = data.get("upLink") or None
        port.addInfo = data.get("addInfo") or None
        port.cabPlacement = data.get("cabPlacement") or None
        port.save()

        return JsonResponse({"success": True})
    except Port.DoesNotExist:
        return JsonResponse({"success": False, "error": "Port not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)

def export_switch_to_excel(request, switch_id):
    switch = Switch.objects.get(pk=switch_id)
    ports = switch.ports.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Коммутатор"

    # Информация о коммутаторе
    ws.append(["Название коммутатора", switch.name])
    ws.append(["Расположение", switch.placement])
    ws.append(["Серийный номер", switch.serialNum])
    ws.append(["Инвентарный номер", switch.inventoryNum])
    ws.append(["MAC-Адрес", switch.macAddr])
    ws.append(["IP-адрес", switch.ip_address])
    ws.append(["Количество портов", switch.ports_count])

    # Пустая строка
    ws.append([])

    # Заголовки таблицы портов
    ws.append([
        "№ порта", "Подключенное устройство", "IP-адрес устройства", "Название устройства",
        "Тип порта", "Состояние", "VLAN", "Up Link",
        "Доп. информация", "Расположение"
    ])

    # Данные портов
    for port in ports:
        ws.append([
            port.number,
            port.connected_device,
            port.ip_address,
            port.device_name,
            port.portType,
            port.status,
            port.vlanNum,
            port.upLink,
            port.addInfo,
            port.cabPlacement,
        ])
    # Заголовки коммутатора жирным
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # Заголовки портов жирным
    for cell in ws[9]:
        cell.font = Font(bold=True)
    # Отправляем файл пользователю
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=switch_{switch.name}.xlsx'
    wb.save(response)
    return response